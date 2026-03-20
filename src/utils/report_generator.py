"""
SamiX Audit Export Engine

This module is responsible for generating professional, portable audit reports.
It provides high-fidelity exports for stakeholders and detailed data sheets 
for operational analysis.

Supported Formats:
1. Branded PDF: A visually rich, supervisor-ready document with KPI charts, 
   verdicts, and violation highlights (powered by ReportLab).
2. Analytics Excel: A multi-sheet workbook containing raw transcripts, 
   granular scores, and cost breakdowns (powered by openpyxl).
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from src.utils.history_manager import AuditSession


# Visual Identity Constants 
# These RGB values match the 'Electric Violet' UI theme.

_INK   = (33,  42,  49)
_VIOLA = (139, 92,  246)
_GREEN = (16,  185, 129)
_RED   = (239, 68,  68)
_AMBER = (245, 158, 11)
_LIGHT = (226, 232, 240)
_SLATE = (100, 116, 135)


class ReportGenerator:
    """
    Orchestrates the creation of audit artifacts.
    
    Transforms the complex AuditSession object into clean, readable 
    documents suitable for performance reviews and compliance archival.
    """

    # PDF Generation (ReportLab) 

    def to_pdf(self, session: "AuditSession") -> bytes:
        """
        Generates a high-quality PDF report.
        
        Args:
            session: The AuditSession record to export.
            
        Returns:
            Raw bytes of the generated PDF document.
        """
        try:
            return self._build_pdf(session)
        except ImportError:
            # Safe fallback if ReportLab is missing in the runtime environment.
            return self._pdf_fallback(session)

    def _build_pdf(self, session: "AuditSession") -> bytes:
        """ Internal logic for constructing the branded PDF layout. """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, HRFlowable,
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

        buf    = io.BytesIO()
        doc    = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=18*mm, rightMargin=18*mm,
            topMargin=18*mm, bottomMargin=18*mm,
        )
        styles = getSampleStyleSheet()

        # Helper to convert RGB tuples to ReportLab Color objects.
        def c(rgb):
            return colors.Color(rgb[0]/255, rgb[1]/255, rgb[2]/255)

        # Style Definitions
        h1 = ParagraphStyle("h1", parent=styles["Normal"],
                             fontSize=20, textColor=c(_VIOLA),
                             fontName="Helvetica-Bold", spaceAfter=4)
        h2 = ParagraphStyle("h2", parent=styles["Normal"],
                             fontSize=13, textColor=c(_INK),
                             fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=10)
        body = ParagraphStyle("body", parent=styles["Normal"],
                                fontSize=9, textColor=c(_INK),
                                leading=14, spaceAfter=2)
        meta = ParagraphStyle("meta", parent=styles["Normal"],
                                fontSize=8, textColor=c(_SLATE), leading=12)
        violation_style = ParagraphStyle("viol", parent=styles["Normal"],
                                          fontSize=8, textColor=c(_RED),
                                          leading=13, leftIndent=10)
        correct_style = ParagraphStyle("corr", parent=styles["Normal"],
                                        fontSize=8, textColor=c(_GREEN),
                                        leading=13, leftIndent=10)

        s = session.scores
        story = []

        # Branding & Header 
        story.append(Paragraph("SamiX · samiksha", h1))
        story.append(Paragraph("Quality Audit Report", h1))
        story.append(HRFlowable(width="100%", color=c(_VIOLA), thickness=1))
        story.append(Spacer(1, 4*mm))

        # Session Metadata Table
        meta_rows = [
            ["File / session",  session.filename],
            ["Stored as",       session.stored_name],
            ["Date / time",     session.upload_time],
            ["Agent",           session.agent_name],
            ["Duration",        f"{session.duration_sec}s"],
            ["Mode",            session.mode.capitalize()],
        ]
        t = Table(meta_rows, colWidths=[50*mm, 120*mm])
        t.setStyle(TableStyle([
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("FONTNAME",     (0,0), (0,-1), "Helvetica-Bold"),
            ("TEXTCOLOR",    (0,0), (0,-1), c(_SLATE)),
            ("TEXTCOLOR",    (1,0), (1,-1), c(_INK)),
            ("ROWBACKGROUNDS",(0,0),(-1,-1), [colors.white, colors.Color(.96,.97,.98)]),
            ("GRID",         (0,0), (-1,-1), 0.25, c(_LIGHT)),
            ("TOPPADDING",   (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ]))
        story.append(t)
        story.append(Spacer(1, 5*mm))

        # Executive Verdict Banner 
        verdict_colour = (
            c(_GREEN) if s.final_score >= 80
            else c(_VIOLA) if s.final_score >= 70
            else c(_AMBER) if s.final_score >= 60
            else c(_RED)
        )
        banner = Table(
            [[
                Paragraph(f"<b>{s.final_score:.0f}/100</b>", ParagraphStyle(
                    "banner_score", parent=styles["Normal"],
                    fontSize=28, textColor=verdict_colour,
                    fontName="Helvetica-Bold"
                )),
                Paragraph(
                    f"<b>{s.verdict}</b><br/>"
                    f"<font size=8 color='#{_SLATE[0]:02x}{_SLATE[1]:02x}{_SLATE[2]:02x}'>"
                    f"Customer sentiment: {s.customer_overall:.1f}/10 &nbsp;·&nbsp; "
                    f"Violations: {session.violations}</font>",
                    body
                ),
            ]],
            colWidths=[40*mm, 130*mm],
        )
        banner.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.Color(.97,.96,.99)),
            ("BOX",        (0,0), (-1,-1), 1, c(_VIOLA)),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",(0,0), (-1,-1), 8),
            ("RIGHTPADDING",(0,0),(-1,-1), 8),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ]))
        story.append(banner)
        story.append(Spacer(1, 5*mm))

        # Performance Breakdown 
        story.append(Paragraph("Dimension Scores", h2))
        dim_data = [["Dimension", "Score /10", "Weight", "Contribution"]] + [
            [name, f"{val:.1f}", f"{int(wt*100)}%", f"{val*wt*10:.1f} pts"]
            for name, val, wt in [
                ("Empathy",        s.empathy,        0.20),
                ("Professionalism", s.professionalism, 0.15),
                ("Compliance",     s.compliance,      0.25),
                ("Resolution",     s.resolution,      0.20),
                ("Communication",  s.communication,   0.05),
                ("Integrity",      s.integrity,       0.15),
            ]
        ]
        dim_data.append(["Phase bonus", f"{s.phase_bonus:+.1f}", "", ""])
        dim_data.append(["FINAL SCORE", f"{s.final_score:.0f}/100", "", ""])

        dt = Table(dim_data, colWidths=[60*mm, 30*mm, 25*mm, 55*mm])
        dt.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), c(_VIOLA)),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, colors.Color(.97,.97,.99)]),
            ("BACKGROUND",  (0,-1), (-1,-1), colors.Color(.96,.97,.98)),
            ("FONTNAME",    (0,-1), (-1,-1), "Helvetica-Bold"),
            ("GRID",        (0,0), (-1,-1), 0.25, c(_LIGHT)),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ]))
        story.append(dt)
        story.append(Spacer(1, 5*mm))

        # AI Analysis 
        if session.summary:
            story.append(Paragraph("AI-Generated Call Summary", h2))
            story.append(Paragraph(session.summary, body))
            story.append(Spacer(1, 4*mm))

        # Critical Failures (Wrong Turns) 
        if session.wrong_turns:
            story.append(Paragraph("Where It Went Wrong", h2))
            for wt in session.wrong_turns:
                story.append(Paragraph(
                    f"<b>Turn {wt.turn_number} · {wt.speaker} · {wt.timestamp}</b>  "
                    f"<font size=8 color='#{_SLATE[0]:02x}{_SLATE[1]:02x}{_SLATE[2]:02x}'>"
                    f"{wt.score_impact}</font>",
                    body
                ))
                story.append(Paragraph(f'Said: "{wt.agent_said}"', meta))
                story.append(Paragraph(f"Wrong: {wt.what_went_wrong}", violation_style))
                story.append(Paragraph(f"Correct: {wt.correct_fact}", correct_style))
                story.append(Paragraph(
                    f"Specific correction: {wt.specific_correction}", body
                ))
                story.append(Paragraph(
                    f"RAG source: {wt.rag_source} · conf {wt.rag_confidence:.2f}",
                    meta
                ))
                story.append(Spacer(1, 3*mm))

        # Finalize and build
        doc.build(story)
        return buf.getvalue()

    @staticmethod
    def _pdf_fallback(session: "AuditSession") -> bytes:
        """ Plain-text alternative used if high-fidelity rendering is unavailable. """
        lines = [
            "SAMIX QUALITY AUDIT REPORT (TEXT FALLBACK)",
            "=" * 50,
            f"File:   {session.filename}",
            f"Date:   {session.upload_time}",
            f"Score:  {session.scores.final_score:.0f}/100",
            f"Verdict:{session.scores.verdict}",
            "",
            session.summary or "",
        ]
        return "\n".join(lines).encode("utf-8")

    # Excel Export (openpyxl) 

    def to_excel(self, session: "AuditSession") -> bytes:
        """
        Exports the audit session history to a multi-sheet Excel workbook.
        
        Sheet Structure:
        - Summary: Metadata and final verdict.
        - Scores: Granular dimensional weights.
        - Violations: RAG-grounded error analysis.
        - Transcript: Complete turn-by-turn log.
        - Cost: API consumption breakdown.
        """
        try:
            return self._build_excel(session)
        except ImportError:
            return b""

    def _build_excel(self, session: "AuditSession") -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        s  = session.scores

        def _head_fill(hex_color: str):
            return PatternFill("solid", fgColor=hex_color)

        def _border():
            thin = Side(style="thin", color="CCCCCC")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def _fmt_sheet(ws, headers: list[str], rows: list[list], col_widths: list[int]):
            """ Helper to apply industrial styling to spreadsheet sheets. """
            ws.append(headers)
            for cell in ws[1]:
                cell.font    = Font(bold=True, color="FFFFFF")
                cell.fill    = _head_fill("1E293B")
                cell.alignment = Alignment(horizontal="center")
                cell.border  = _border()
            for row in rows:
                ws.append(row)
            for row_cells in ws.iter_rows(min_row=2):
                for cell in row_cells:
                    cell.border = _border()
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # Sheet 1: Executive Summary 
        ws1 = wb.active
        ws1.title = "Summary"
        _fmt_sheet(ws1,
            ["Field", "Value"],
            [
                ["Filename",          session.filename],
                ["Date",              session.upload_time],
                ["Agent",             session.agent_name],
                ["Final QA score",    f"{s.final_score:.0f}/100"],
                ["Verdict",           s.verdict],
                ["Violations",        session.violations],
                ["API cost (USD)",    f"${session.cost_usd:.5f}"],
            ],
            [30, 80],
        )

        # Sheet 2: Dimension Scores 
        ws2 = wb.create_sheet("Scores")
        _fmt_sheet(ws2,
            ["Dimension", "Score /10", "Weight %", "Contribution pts"],
            [
                ["Empathy",        s.empathy,        20, round(s.empathy * 0.20 * 10, 2)],
                ["Professionalism", s.professionalism, 15, round(s.professionalism * 0.15 * 10, 2)],
                ["Compliance",     s.compliance,      25, round(s.compliance * 0.25 * 10, 2)],
                ["Resolution",     s.resolution,      20, round(s.resolution * 0.20 * 10, 2)],
                ["Communication",  s.communication,    5, round(s.communication * 0.05 * 10, 2)],
                ["Integrity",      s.integrity,       15, round(s.integrity * 0.15 * 10, 2)],
                ["FINAL SCORE",    s.final_score,    100, s.final_score],
            ],
            [25, 14, 12, 18],
        )

        # Sheet 3: Violations 
        ws3 = wb.create_sheet("Violations")
        _fmt_sheet(ws3,
            ["Turn", "Speaker", "Timestamp", "Agent said",
             "What went wrong", "Correct fact", "RAG source",
             "RAG conf", "Score impact", "Specific correction"],
            [
                [wt.turn_number, wt.speaker, wt.timestamp, wt.agent_said,
                 wt.what_went_wrong, wt.correct_fact, wt.rag_source,
                 wt.rag_confidence, wt.score_impact, wt.specific_correction]
                for wt in session.wrong_turns
            ],
            [6, 10, 10, 30, 35, 35, 25, 8, 18, 40],
        )

        # Sheet 4: Transcript 
        ws4 = wb.create_sheet("Transcript")
        _fmt_sheet(ws4,
            ["Turn", "Speaker", "Timestamp", "Text", "Sentiment 0-10"],
            [
                [t.turn, t.speaker, t.timestamp, t.text,
                 s.customer_sentiment[t.turn - 1]
                 if t.turn <= len(s.customer_sentiment) else ""]
                for t in session.transcript
            ],
            [6, 12, 10, 80, 14],
        )
        # Colour agent rows violet tint
        for row in ws4.iter_rows(min_row=2):
            if row[1].value == "AGENT":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="F3F0FF")

        # Sheet 5: Cost 
        ws5 = wb.create_sheet("Cost")
        _fmt_sheet(ws5,
            ["Item", "Value"],
            [
                ["Tokens used",       session.token_count],
                ["API cost (USD)",    round(session.cost_usd, 5)],
                ["Revenue / audit",   5.0],
                ["Profit / audit",    round(5.0 - session.cost_usd, 4)],
                ["Margin %",          round((5.0 - session.cost_usd) / 5.0 * 100, 2)],
            ],
            [25, 18],
        )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
